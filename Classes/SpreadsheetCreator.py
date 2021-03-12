import os

def getTabName():
    import datetime
    current_time = datetime.datetime.now()
    monthlyFolderTitle = current_time.strftime('%m%d%y')
    return monthlyFolderTitle

from SetupTool import getMonthlyFolderTitle

class SpreadsheetCreator:
    def __init__(self, workbook):
        self.workbook = workbook
        self.tabName = getTabName()
        self.directory = "Recons/"+getMonthlyFolderTitle()



    def doubleAccountReconcilliationformat(self, account_Balance):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import datetime
        #print(os.getcwd())
        os.chdir(self.directory)
        wb = Workbook()
        wb.create_sheet(self.tabName)
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        sheet = wb.active
        #Merge the cells
        sheet.merge_cells("A1:J1")
        sheet.merge_cells("A2:J2")
        sheet.merge_cells("A4:B4")
        sheet.merge_cells("F4:G4")
        #Allign the cells
        fontObj1 = Font(name='Garamond',bold=True)
        sheet['A1'].font = fontObj1
        sheet['A2'].font = fontObj1
        sheet['A1'].alignment = Alignment(horizontal="center")
        sheet['A2'].alignment = Alignment(horizontal="center")
        #Enter in information
        sheet['A1'] = str(self.workbook)
        sheet['A2'] = datetime.date.today()
        sheet['A4'] = 'Business Fundamentals Chk - 9387'
        sheet['E4'] = 'AccountBalanceVariable'
        sheet['F4'] = 'Quickbooks Account Balance - 100500'
        sheet['A6'] = 'DATE'
        sheet['B6'] = 'ADJUSTMENTS:'
        sheet['C6'] = 'CLEAR DATE'
        sheet['D6'] = 'INITIALS'
        sheet['E6'] = 'AMOUNT'
        sheet['F6'] = 'DATE'
        sheet['G6'] = 'ADJUSTMENTS:'
        sheet['H6'] = 'CLEAR DATE'
        sheet['I6'] = 'INITIALS'
        sheet['J6'] = 'AMOUNT'
        sheet['A16'] = 'Adjusted Account Balance'
        sheet['E16'] = '=SUM(E4:E15)'
        sheet['F16'] = 'Adjusted GL Balance'
        sheet['J16'] = '=sum(J8:J14)'
        sheet['H18'] = 'VARIANCE'
        sheet['J18'] = '=E16-J16'
        sheet['J4'] = account_Balance
        workbookName = str(self.workbook).replace("/", "")
        workbookName.strip("/")
        #os.chdir(self.directory)
        #print(self.directory)
        wb.save(self.workbook)

    def singleAccountReconcilliationformat(self,account_Balance):
       from openpyxl import Workbook
       from openpyxl.styles import Font, Alignment
       import datetime
       os.chdir(self.directory)
       wb = Workbook()
       wb.create_sheet(self.tabName)
       wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
       sheet = wb.active
       # Merge the cells
       sheet.merge_cells("A1:L1")
       sheet.merge_cells("A2:L2")
       sheet.merge_cells("A5:H5")
       sheet.merge_cells("I5:J5")
       sheet.merge_cells("B7:H7")
       sheet.merge_cells("I9:J9")
       sheet.merge_cells("I10:J10")
       sheet.merge_cells("I11:J11")
       sheet.merge_cells("I12:J12")
       sheet.merge_cells("I13:J13")
       sheet.merge_cells("I14:J14")
       sheet.merge_cells("I15:J15")
       sheet.merge_cells("I16:J16")
       sheet.merge_cells("B9:H9")
       sheet.merge_cells("B10:H10")
       sheet.merge_cells("B11:H11")
       sheet.merge_cells("B12:H12")
       sheet.merge_cells("B13:H13")
       sheet.merge_cells("B14:H14")
       sheet.merge_cells("B15:H15")
       sheet.merge_cells("B16:H16")
       sheet.merge_cells("I19:J19")
       # Allign the cells
       fontObj1 = Font(name='Garamond', bold=True)
       sheet['A1'].font = fontObj1
       sheet['A2'].font = fontObj1
       sheet['A1'].alignment = Alignment(horizontal="center")
       sheet['A2'].alignment = Alignment(horizontal="center")
       sheet['A5'].alignment = Alignment(horizontal="right")
       # Enter in information
       sheet['A1'] = str(self.workbook)
       sheet['A2'] = datetime.date.today()
       sheet['A5'] = 'GL Balance:'
       sheet['A7'] = 'Date:'
       sheet['B7'] = 'Description'
       sheet['I7'] = 'Amount'
       sheet['H19'] = 'Variance'
       sheet['I19'] = '=I5-SUM(I9:J16)'
       sheet['J4'] = account_Balance
       workbookName = str(self.workbook).replace("/", "")



       wb.save(workbookName)
