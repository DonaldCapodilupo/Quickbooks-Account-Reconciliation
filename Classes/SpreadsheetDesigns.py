def getTabName():
    import datetime
    current_time = datetime.datetime.now()
    monthlyFolderTitle = current_time.strftime('%m%d%y')
    return monthlyFolderTitle

from SetupTool import getMonthlyFolderTitle

class SpreadsheetDesigner:
    def __init__(self, workbook):
        self.workbook = workbook
        self.tabName = getTabName()
        self.directory = "Recons/"+getMonthlyFolderTitle()


    def moveAndCopyWorksheet(self):
        import os
        os.chdir(self.directory)

        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        try:
            wb = load_workbook(self.workbook)
            previous_worksheet = wb.worksheets[0]
            new_worksheet = wb.create_sheet(self.tabName)
            for row in previous_worksheet:
                for cell in row:
                    new_worksheet[cell.coordinate].value = cell.value
            wb.save(self.workbook)
        except InvalidFileException:
            print("Not an excel file extension - "+ self.workbook)



    def insertNewAccountBalances(self, accountBalance):
        from openpyxl import load_workbook
        wb = load_workbook(self.workbook)
        ws = wb[self.tabName]
        ws["J4"] = accountBalance
        wb.save(self.workbook)

